#!/usr/bin/env python3
"""
Import real data from XLSX master spreadsheet into DTL database.

Usage (run from backend/ directory):
    python tools/import_xlsx.py path/to/master.xlsx [--clear]

  --clear   Wipe all transactional data before import (keeps users, sites, tariffs)

The script is idempotent when run without --clear only if records don't already exist.
With --clear it first deletes all rows from transactional tables.
"""
import sys
import os
import argparse
from datetime import datetime, date

# ── allow running from backend/ dir ──────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
except ImportError:
    print("Установите openpyxl: pip install openpyxl"); sys.exit(1)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("Установите psycopg2: pip install psycopg2-binary"); sys.exit(1)

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    print("Не найдена переменная DATABASE_URL (проверьте .env)"); sys.exit(1)


def connect():
    return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)


def exec_sql(cur, sql, params=None):
    cur.execute(sql, params)


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def normalize(s):
    """Strip and normalize whitespace."""
    if s is None:
        return None
    return " ".join(str(s).split()).strip()


# ── Reference data helpers ────────────────────────────────────────────────────

def ensure_client(cur, name):
    name = normalize(name)
    if not name:
        return None
    cur.execute("SELECT id FROM clients WHERE LOWER(name) = LOWER(%s)", (name,))
    row = cur.fetchone()
    if row:
        return row["id"]
    cur.execute("INSERT INTO clients (name) VALUES (%s) RETURNING id", (name,))
    return cur.fetchone()["id"]


def ensure_carrier(cur, name):
    name = normalize(name)
    if not name:
        return None
    cur.execute("SELECT id FROM carriers WHERE LOWER(name) = LOWER(%s)", (name,))
    row = cur.fetchone()
    if row:
        return row["id"]
    cur.execute("INSERT INTO carriers (name) VALUES (%s) RETURNING id", (name,))
    return cur.fetchone()["id"]


def ensure_supplier(cur, name):
    name = normalize(name)
    if not name:
        return None
    cur.execute("SELECT id FROM suppliers WHERE LOWER(name) = LOWER(%s)", (name,))
    row = cur.fetchone()
    if row:
        return row["id"]
    cur.execute("INSERT INTO suppliers (name) VALUES (%s) RETURNING id", (name,))
    return cur.fetchone()["id"]


def ensure_truck(cur, name):
    name = normalize(name)
    if not name:
        return None
    cur.execute("SELECT id FROM trucks WHERE LOWER(name) = LOWER(%s)", (name,))
    row = cur.fetchone()
    if row:
        return row["id"]
    # Create as active DTL truck
    cur.execute("INSERT INTO trucks (name, owner, status) VALUES (%s, 'DTL', 'active') RETURNING id", (name,))
    return cur.fetchone()["id"]


# ── Clear test data ───────────────────────────────────────────────────────────

CLEAR_TABLES = [
    "hire_deliveries",
    "income_records",
    "company_expenses",
    "fleet_expenses",
    "debt_records",
    "fuel_dispatches",
    "fuel_receipts",
    "orders",
    "order_sites",
    "cash_to_artem",
    "fuel_advances",
    "monthly_reconciliations",
    "balance_entries",
    "audit_log",
]

CLEAR_REF_TABLES = [
    "carriers",
    "suppliers",
    "clients",
]

TRUCK_NAMES_TO_KEEP = ["Шахман-1", "Шахман-2", "Шахман-3", "Сайгак", "Фусо", "MAN", "Кантер", "Скания"]


def clear_data(cur):
    print("⚠️  Удаляю тестовые данные...")
    for t in CLEAR_TABLES:
        cur.execute(f"DELETE FROM {t}")
        print(f"  ✓ {t}")
    for t in CLEAR_REF_TABLES:
        cur.execute(f"DELETE FROM {t}")
        print(f"  ✓ {t}")
    # Remove test trucks (keep seed trucks by name)
    placeholders = ",".join(["%s"] * len(TRUCK_NAMES_TO_KEEP))
    cur.execute(f"DELETE FROM trucks WHERE name NOT IN ({placeholders})", TRUCK_NAMES_TO_KEEP)
    print(f"  ✓ trucks (удалены тестовые)")
    print("Данные очищены.")


# ── Import sheets ─────────────────────────────────────────────────────────────

def import_hire(ws, cur):
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows[1:] if r[0] is not None and r[2] is not None]
    inserted = 0
    for r in data:
        d = parse_date(r[0])
        client_id = ensure_client(cur, r[2])
        supplier_id = ensure_supplier(cur, r[3]) if r[3] else None
        carrier_id = ensure_carrier(cur, r[4]) if r[4] else None
        volume = float(r[5]) if r[5] else None
        price_client = float(r[6]) if r[6] else None
        price_supplier = float(r[7]) if r[7] else None
        price_carrier = float(r[8]) if r[8] else None
        amount_client = float(r[9]) if r[9] else None
        amount_supplier = float(r[10]) if r[10] else None
        amount_carrier = float(r[11]) if r[11] else None
        margin = float(r[12]) if r[12] else None
        comment = normalize(r[14]) if len(r) > 14 else None

        cur.execute("""
            INSERT INTO hire_deliveries
                (delivery_at, client_id, supplier_id, carrier_id,
                 volume_liters, price_client, price_supplier, price_carrier,
                 amount_client, amount_supplier, amount_carrier, margin, comment)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (d, client_id, supplier_id, carrier_id,
              volume, price_client, price_supplier, price_carrier,
              amount_client, amount_supplier, amount_carrier, margin, comment))
        inserted += 1
    print(f"  ✓ Найм: {inserted} строк")


def import_debts(ws, cur):
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows[1:] if r[0] is not None and r[3] is not None]
    inserted = 0
    for r in data:
        d = parse_date(r[0])
        debtor = normalize(r[3])
        amount = float(r[4]) if r[4] else 0
        dtype = normalize(r[5]) or "ДОЛГ"
        comment = normalize(r[6]) if len(r) > 6 else None

        cur.execute("""
            INSERT INTO debt_records (recorded_at, debtor, amount, type, comment)
            VALUES (%s, %s, %s, %s, %s)
        """, (d, debtor, amount, dtype, comment))
        inserted += 1
    print(f"  ✓ Долги: {inserted} строк")


def import_income(ws, cur):
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows[1:] if r[0] is not None and r[2] is not None]
    inserted = 0
    for r in data:
        d = parse_date(r[0])
        client_id = ensure_client(cur, r[2])
        amount = float(r[3]) if r[3] else 0
        volume = float(r[4]) if r[4] else None
        comment = normalize(r[5]) if len(r) > 5 else None

        cur.execute("""
            INSERT INTO income_records (income_at, client_id, amount, volume, comment)
            VALUES (%s, %s, %s, %s, %s)
        """, (d, client_id, amount, volume, comment))
        inserted += 1
    print(f"  ✓ Доходы: {inserted} строк")


def import_expenses(ws, cur):
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows[1:] if r[0] is not None and r[2] is not None]
    inserted = 0
    for r in data:
        d = parse_date(r[0])
        if d is None:
            continue
        category = normalize(r[2]) or "Прочие"
        amount = float(r[3]) if r[3] else 0
        comment = normalize(r[4]) if len(r) > 4 else None

        cur.execute("""
            INSERT INTO company_expenses (expense_at, category, amount, comment)
            VALUES (%s, %s, %s, %s)
        """, (d, category, amount, comment))
        inserted += 1
    print(f"  ✓ Общие расходы: {inserted} строк")


# Category mapping from spreadsheet column to DB category name
MACHINE_EXPENSE_COLS = {
    3: "Резина/расходники",
    4: "ТО",
    5: "Прочие расходы",
    6: "Ремонт",
    7: "Зарплата",
    8: "Топливо",
    10: "Парковка/база",
    11: "Налоги/штрафы",
    12: "Страховка",
    13: "Кредит/лизинг",
}


def import_fleet_expenses(ws, cur):
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows[1:] if r[0] is not None and r[2] is not None]
    inserted = 0
    for r in data:
        d = parse_date(r[0])
        if d is None:
            continue
        truck_id = ensure_truck(cur, r[2])
        trips = int(r[9]) if len(r) > 9 and r[9] else None
        revenue = float(r[14]) if len(r) > 14 and r[14] else None
        comment = normalize(r[15]) if len(r) > 15 else None

        for col_idx, cat_name in MACHINE_EXPENSE_COLS.items():
            if len(r) > col_idx and r[col_idx]:
                amount = float(r[col_idx])
                cur.execute("""
                    INSERT INTO fleet_expenses
                        (truck_id, expense_at, category, amount, trips, revenue, comment)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (truck_id, d, cat_name, amount,
                      trips if cat_name == "Зарплата" else None,
                      revenue if cat_name == "Зарплата" else None,
                      comment))
                inserted += 1
    print(f"  ✓ Расходы по машинам: {inserted} строк")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import DTL data from XLSX")
    parser.add_argument("xlsx", help="Path to master XLSX file")
    parser.add_argument("--clear", action="store_true",
                        help="Delete all test/transactional data before import")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"Файл не найден: {args.xlsx}"); sys.exit(1)

    print(f"Загружаю {args.xlsx}...")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    required = {"Найм", "Долги", "Доходы", "Общие расходы", "Машины_месяц"}
    missing = required - set(wb.sheetnames)
    if missing:
        print(f"В файле нет листов: {missing}"); sys.exit(1)

    conn = connect()
    cur = conn.cursor()

    try:
        if args.clear:
            clear_data(cur)
            conn.commit()

        print("Импортирую данные...")
        import_hire(wb["Найм"], cur)
        import_debts(wb["Долги"], cur)
        import_income(wb["Доходы"], cur)
        import_expenses(wb["Общие расходы"], cur)
        import_fleet_expenses(wb["Машины_месяц"], cur)

        conn.commit()
        print("\n✅ Импорт завершён успешно!")

    except Exception as e:
        conn.rollback()
        print(f"\n❌ Ошибка: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
